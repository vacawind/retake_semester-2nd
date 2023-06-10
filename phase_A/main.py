import os
import time
# import shutil
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side 
from tqdm import tqdm

# function 數B數C分類
def classify_math(row):
    if "園藝" in row["班級"] or "食品" in row["班級"]:
        return "數學B"
    else:
        return "數學C"
    
# function 基本電學分類
def classify_basic_electric(row):
    if "電機" in row["班級"]:
        return "電機部必3"
    
    elif "汽車" in row["班級"] or "機車" in row["班級"]:
        return "汽機部必2"
    
    else:
        return "生機校必2"

# function 分科課程生成 
def gernerate_course(title, subject, grade, current_dir, raw_path):
    df = pd.read_excel(raw_path) # 讀 raw.xlsx 生成 df
    
    # df科目分類整理
    df.columns = df.iloc[0]
    df = df.iloc[1:,].reindex()
    df = df.loc[df['科目'] == subject]
    df.reset_index(drop=True, inplace=True)
    
    # 過濾出數B與數C, 並生成excel
    if subject == '數學Ⅰ':
        df["科目類別"] = df.apply(classify_math, axis=1)
        math_subjects = ['數學B', '數學C']
        for math in math_subjects:
            if grade == '高一科目':
                df_math = df[(df["科目類別"] == math) & (df["必選修類別"] == '部定必修')]
            elif grade == '高二科目':
                df_math = df[(df["科目類別"] == math) & (df["必選修類別"] == '校訂必修')]
                
            # df_math = df[df["科目類別"] == math]
            students_num = len(df_math.index)
            new_title = title[0:2] + math + title[4:]
            filename = f'{new_title}-{students_num}.xlsx'
            df_math.to_excel(os.path.join(current_dir, f'!!{grade}', filename))
        return
    
    # 過濾出基本電學(電機,生機,汽機車), 並生成excel
    if subject == '基本電學Ⅰ':
        df["科目類別"] = df.apply(classify_basic_electric, axis=1)
        electric_subjects = ['電機部必3', '汽機部必2', '生機校必2']
        for electric in electric_subjects:
            df_electric = df[df["科目類別"] == electric]
            students_num = len(df_electric.index)
            new_title = f'({electric})' + title[0:6] 
            filename = f'{new_title}-{students_num}.xlsx'
            df_electric.to_excel(os.path.join(current_dir, f'!!{grade}', filename))
        return
    
    # 科目人數excel檔名生成  
    students_num = len(df.index)
    filename = f'{title}-{students_num}.xlsx'
    df.to_excel(os.path.join(current_dir, f'!!{grade}', filename))

# function excel格式整理 
def reform_excel(f_path):
    wb = openpyxl.load_workbook(f_path)
    ws = wb.worksheets[0]
    # 刪除第一欄
    ws.delete_cols(1)
    
    # 刪除『科目類別』欄位
    subject_category_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == '科目類別':
            subject_category_col = col
            break
    if subject_category_col is not None:
        ws.delete_cols(subject_category_col)
    
    # 在備註欄後方加入'是否參加'欄位
    next_cell = None
    for cell in ws[1]:
        if cell.value == '備註':
            next_cell = cell.offset(row=0, column=1)    
            break
    ws[next_cell.coordinate] = '是否參加'
    
    # 字體改為標楷體
    # 儲存格資料置中
    font = Font(name='標楷體')
    alignment = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
            cell.alignment = alignment
    
    # 設定完篩選範圍        
    ws.auto_filter.ref = f'A1:{next_cell.coordinate}'
    
    # 刪除框線   
    no_border = Border(left=Side(style='none'), 
                right=Side(style='none'), 
                top=Side(style='none'), 
                bottom=Side(style='none'))
    for row in ws.rows:
        for cell in row:
            cell.border = no_border
            
    wb.save(f_path) 
    
# 桌面根目錄
current_dir = os.path.dirname(__file__)
# raw.xlsx 絕對路徑
raw_path = os.path.join(current_dir, 'raw.xlsx')   

# 年級科目清單
grade_subject = ['高一科目',
                 '高二科目',
                 '高三科目',
                 ]

# 設定開始執行時間點
start_time = time.time()
print('******* 下學期重補修課程生成作業 phaseA ******* ')

# 從年級科目清單開始建立課程生成
for grade in grade_subject:
    print(f'\n正在處理的年級為：{grade}')
    subjects = [] 
    with open(os.path.join(current_dir, f'{grade}.csv'), 'r', encoding='utf-8') as f:  
        # 從高一/高二/高三科目.csv分類出『標題』『科目』
        for line in f:
            title, subject = line.strip().split(',')
            subjects.append([title, subject])
    
    # 分科課程生成
    pbar = tqdm(subjects, desc='重補修課程生成 Processing') # 將 subjects 引進進度條模組       
    for sub in pbar:
        title = sub[0]
        subject = sub[1]
        gernerate_course(title, subject, grade, current_dir, raw_path)
        time.sleep(0.2)
            
    # excel格式整理   
    pbar = tqdm(os.listdir(os.path.join(current_dir, f'!!{grade}')), desc='Excel格式重整 Processing')
    for file in pbar:
        if file =='.DS_Store' or file == file.startswith('~$') or file == '.gitkeep':
            continue
        file_path = os.path.join(current_dir, f'!!{grade}', file)
        reform_excel(file_path)
        time.sleep(0.2)


# # 進行壓縮檔案處理
# #資料夾列表
# folder_paths = [os.path.join(current_dir, grade) for grade in grade_subject]
# print(folder_paths)
# # 壓縮後的檔案路徑和名稱
# archive_path = os.path.join(current_dir, 'project')
# #壓縮資料夾
# shutil.make_archive(archive_path, 'zip', *folder_paths)

# 設定結束執行時間點    
end_time = time.time()
# 計算共運行時間   
total_time = end_time - start_time
print('\n******* 作業結束 ******* ')
print(f"共耗時：{total_time} 秒")
# 結束運行時螢幕維持3秒後關閉powershell 
time.sleep(3)