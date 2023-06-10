import os
import time
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
from tqdm import tqdm 

# 重補修科目班級分析 
def df_analysis(subject_file_path, department_class):
    df = pd.read_excel(subject_file_path)
    df = df[df['班級'] == department_class]
    
    if df.empty: 
        return None, None, None, None, None
    else:   
        class_name = df['班級']
        student_classId = df['座號']
        student_id = df['學號']
        student_name = df['姓名']
        grade_year = df['學年']
        grade = df['年級']
        grade_section = df['學期']
        retake_subject = df['科目']
        course_classify = df['必選修類別']
        course_points = df['學分']
        
        analysis_datum = [class_name, student_classId, student_id, student_name, grade_year, grade, grade_section, retake_subject]
        subject = retake_subject.iloc[0]
        points = course_points.iloc[0]
        classify = course_classify.iloc[0]
        df_rows = df.shape[0] # df row number
        
        return analysis_datum, subject, points, classify, df_rows
    

# 當前目錄
current_dir = os.path.dirname(__file__)

# 重補修願調查表_範本.xlsx 絕對路徑
temp_file = os.path.join(current_dir, '重補修願調查表_範本.xlsx')

# 年級班級清單
grade_class = ['高一班級',
                '高二班級',
                '高三班級',
                ]

# 設定開始執行時間點
start_time = time.time()
print('******* 下學期重補修課程生成作業 phaseB ******* ')

# 年期班級迴圈, 高一班級 => 高二班級 => 高三班級                   
for grade in grade_class:
    print(f'\n正在處理的年級為：{grade}')
    subjects = []
    subjects_path = os.path.join(current_dir, '重補修科目', grade[0:2] + '科目')
    for current_dir, dirs, files in os.walk(subjects_path):
        subjects = files
        
    current_dir = os.path.dirname(__file__) #不加這行 current_dir 會跑掉
    
    # 打開各班級名稱
    department_classes = []
    with open(os.path.join(current_dir, f'{grade}.csv'), 'r', encoding='utf-8') as f:
        for line in f:
            department_classes.append(line.strip())
            
    # 從班級開始進行迴圈,ex.電機一甲 => 電機一乙 => ...
    pbar = tqdm(department_classes, desc='重補修意願調查表生成 Processing') # 將 department_classes 引進進度條模組
    for department_class in pbar:
        wb = openpyxl.load_workbook(temp_file)
        ws = wb.active
        
        # excel重補修班級名稱
        ws['A14'].value = f'班級：{department_class}'
        # 設定開始進行輸出excel資料的row
        row_startIndex = 16
        
        # 進行重補修科目迴圈,ex. 一上國語文 => 一上國文精讀 => ...
        
        for subject in subjects:
            if subject =='.DS_Store' or subject.startswith('~$') or subject == '.gitkeep':
                continue
            
            # 重補修科目分析出各班學生名單
            subject_file_path = os.path.join(subjects_path, subject)
            result = df_analysis(subject_file_path, department_class)
            
            if result[0] is None: continue
            
            analysis_datum, subject, points, classify, df_rows = result
        
            # excel第一列科目學分費用
            ws['A' + str(row_startIndex)].value = f'科目：{subject}    學分：{int(points)}({classify})    費用：{240*int(points)}元'
            
            # excel合併第一列科目學分費用
            ws.merge_cells(f'A{str(row_startIndex)}:K{str(row_startIndex)}')
            
            heading = [
                '班級名稱',
                '座號',
                '學號',
                '姓名',
                '學年',
                '年級',
                '學期',
                '不及格科目',
                '簽名確認',
                '是否參加',
                '備註',
            ]
            
            # excel標頭欄位
            for i, head in enumerate(heading):
                ws.cell(row=row_startIndex+1, column=i+1).value = head
            
            # excel重補修科目學生名單
            for j, analysis_data in enumerate(analysis_datum):
                for i, data in enumerate(analysis_data): 
                    ws.cell(row=row_startIndex+2+i, column=j+1).value = data   
            
            font = Font(size=10, name='標楷體')
            border = Border(
                            left=Side(style='thin', color='000000'),
                            right=Side(style='thin', color='000000'),
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thin', color='000000')
                        )
            alignment = Alignment(horizontal='center', vertical='center')
            for row in ws.iter_rows(min_row=row_startIndex, max_row=row_startIndex+1+df_rows, min_col=1, max_col=11):
                for cell in row:
                    cell.font = font
                    cell.border = border
                    cell.alignment = alignment
                    
            for i in range(row_startIndex, row_startIndex+1+df_rows+1):
                ws.row_dimensions[i].height = 20
            
            # 重新調整下一個excel資料輸出row
            next_gap = 4
            row_startIndex = row_startIndex + df_rows + next_gap
        
        #將各班整理後的excel調查表儲存在各個班級資料夾中       
        wb.save(os.path.join(current_dir, f'!!{grade}', f'{department_class}-調查表.xlsx'))


# 設定結束執行時間點    
end_time = time.time()
# 計算共運行時間   
total_time = end_time - start_time
print('\n******* 作業結束 ******* ')
print(f"共耗時：{total_time} 秒")
# 結束運行時螢幕維持3秒後關閉powershell 
time.sleep(3)